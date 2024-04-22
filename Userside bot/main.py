from telegram import Update, InputFile
from telegram.ext import Updater, CommandHandler, MessageHandler, filters, ContextTypes, Application
from telegram.ext.filters import MessageFilter
import openpyxl
import helper_funcs
import psycopg2

# VARIABLES ###############################################################################

# Connect to database
conn = psycopg2.connect(database = "Hospital Database (Sadra Hosseini)", 
                        user = "postgres", 
                        host= 'a2ba86d2-669b-4bf8-ab7d-1b63a3e1f1db.hsvc.ir',
                        password = "KzPRmunw4j9hCdlkmXIpOkEzhenL3Jvh",
                        port = 30500)
print('App connected to database!')
cur = conn.cursor()

wb_clinics = openpyxl.load_workbook('Userside bot\\clinics.xlsx')
sheet_clinics = wb_clinics.active

clinics_dict = {}
doctors_dict = {}


############################################################################################

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text('''
    با استفاده از دستور /visit می‌توانید به راحتی نوبت پزشک موردنظر خود را تهیه کنید.
    ''')

async def visit_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:    
    # Reading clinics.xlsx file
    for column in range(sheet_clinics.max_column):
        column_obj = sheet_clinics.cell(row = 1, column = column+1)
        clinics_dict[column_obj.value] = []
        for row in range(sheet_clinics.max_row-1):
            row_obj = sheet_clinics.cell(row = row+2, column = column+1)
            if row_obj.value != None:
                clinics_dict[column_obj.value].append(row_obj.value)
    
    
    # Get doctors table from the database
    cur.execute('SELECT * FROM public.doctors')
    doctors = cur.fetchall()

    for row_index in range(len(doctors)):
        row_lst = list(doctors[row_index])
        doctors_dict[row_lst[0]] = [row_lst[1],row_lst[2],row_lst[3],row_lst[4]]
    
    
    for key in doctors_dict.keys():
        if len(doctors_dict[key][3]) > 3:
            if ',' in list(doctors_dict[key][3]):
                doctors_dict[key][3] = doctors_dict[key][3].split(',')
            elif '،' in list(doctors_dict[key][3]):
                doctors_dict[key][3] = doctors_dict[key][3].split('،')
            else:
                print('Shifts are not correct in doctors database.')
        else:
            shift_lst = []
            shift_lst.append(doctors_dict[key][3])
            doctors_dict[key][3] = shift_lst

    # Get times table from the database
    cur.execute('SELECT * FROM public.times')
    context.user_data['times'] = cur.fetchall()
    
    
    await update.message.reply_text('🩺 فهرست کلینیک‌ها' + f'\n\n{'\n'.join(helper_funcs.ordered_text(list(clinics_dict.keys())))}\n\n' + '✅ شماره کلینیک موردنظر خود را وارد کنید.')

    context.user_data['level'] = 1
    
async def visit_process(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('level') == 1: # Level 1: Select the clinic
        try:
            user_clinic = int(update.message.text)
        except:
            await update.message.reply_text('❌ پیام اشتباه! لطفا شماره کلینیک موردنظر خود را وارد کنید.')
        if user_clinic-1 in range(len(clinics_dict.keys())):
            selected_clinic = list(clinics_dict.keys())[user_clinic-1]
            await update.message.reply_text(f'بخش‌های {selected_clinic}\n\n' + '\n'.join(helper_funcs.ordered_text(clinics_dict[selected_clinic])) + '\n\n ✅ شماره بخش موردنظر خود را وارد کنید.')
            context.user_data['level'] = 2
            context.user_data['user_choice_level_1'] = selected_clinic
        else:
            await update.message.reply_text('❌ پیام اشتباه! لطفا در وارد کردن شماره کلینیک موردنظر دقت فرمایید.') 

    elif context.user_data.get('level') == 2: # Level 2: Select the section of clinic
        try:
            user_section = int(update.message.text)
        except:
            await update.message.reply_text('❌ پیام اشتباه! لطفا شماره بخش موردنظر خود را وارد کنید.')
        if user_section-1 in range(len(clinics_dict[context.user_data['user_choice_level_1']])):
            selected_section = clinics_dict[context.user_data['user_choice_level_1']][user_section-1]
            doctors = helper_funcs.find_doctors(selected_section,doctors_dict)
            await update.message.reply_text('👨‍⚕️👩‍⚕️ فهرست پزشکان\n\n' + helper_funcs.show_doctor_results(doctors,doctors_dict) + '\n\n ✅ شماره پزشک و شیفت موردنظر خود را مطابق نمونه وارد کنید.(نمونه متن ارسالی: 2/صبح)')
            context.user_data['level'] = 3
            context.user_data['user_choice_level_2'] = selected_section
            context.user_data['user_doctors'] = doctors
        else:
            await update.message.reply_text('❌ پیام اشتباه! لطفا در وارد کردن شماره بخش موردنظر دقت فرمایید.') 
    
    elif context.user_data.get('level') == 3: # Level 3: Select the doctor
        try:
            user_doctorandshift = update.message.text.split('/')
        except:
            await update.message.reply_text('❌ متن ارسالی را مطابق نمونه داده شده وارد کنید.')
        if int(user_doctorandshift[0])-1 in range(len(context.user_data['user_doctors'])):
            selected_doctor = context.user_data['user_doctors'][int(user_doctorandshift[0])-1]
            if user_doctorandshift[1] in doctors_dict[selected_doctor][3]:
                times = []
                for item in context.user_data['times']:
                    if (list(item)[0] == selected_doctor) and (list(item)[2] == user_doctorandshift[1]):
                        times.append(list(item))
                await update.message.reply_text('نوبت های موجود\n\n'+ helper_funcs.show_times_results(times) +'\n\n✅ شماره نوبت مورد نظر خود را وارد کنید.' )
                context.user_data['level'] = 4
                context.user_data['user_choice_level_3'] = times
            else:
                await update.message.reply_text('☹️ نوبت موردنظر شما در حال حاضر موجود نمی‌باشد.')
        else:
            await update.message.reply_text('❌ پیام اشتباه! لطفا در تایپ کردن شماره پزشک موردنظرتان دقت فرمایید.')
                          
    
    elif context.user_data.get('level') == 4: # Level 4: Select the visit time
        pass
        # user_doctorandshift = update.message.text.split('/')
        # if int(user_doctorandshift[0])-1 in range(len(context.user_data['user_doctors'])):
        #     selected_doctor = context.user_data['user_doctors'][int(user_doctorandshift[0])-1]
        #     try:
        #         if user_doctorandshift[1] in doctors_dict[selected_doctor][3]:
        #             await update.message.reply_text('''
        #             📞🔢 جهت تایید نهایی نوبت، شماره تلفن و کدملی خود را با اعداد انگلیسی وارد کنید.
        #             (فرمت ارسال: شماره تلفن/کدملی)
        #             ''')
        #             user_doctorandshift[0] = selected_doctor
        #             context.user_data['level'] = 5
        #             context.user_data['user_choice_level_3'] = user_doctorandshift
        #         else:
        #             await update.message.reply_text('☹️ شیفت موردنظر شما برای این پزشک موجود نمی‌باشد.')
        #     except:
        #         await update.message.reply_text('❌ پیام اشتباه! شیفت پزشک موردنظر خود را وارد کنید.')
        # else:
        #     await update.message.reply_text('❌ پیام اشتباه! لطفا در تایپ کردن شماره پزشک موردنظرتان دقت فرمایید.') 
    
    elif context.user_data.get('level') == 5: # Level 5: Get personal info from user
        user_personal_info = update.message.text.split('/')
        if (str(user_personal_info[0])[0:2] == '09' and len(str(user_personal_info[0])) == 11) and len(str(user_personal_info[1])) == 10:
            await update.message.reply_text(f'درخواست شما جهت تهیه نوبت با این اطلاعات ثبت شد.\nنام پزشک:{context.user_data['user_choice_level_3'][0]}\nشیفت:{context.user_data['user_choice_level_3'][1]}\nشماره تلفن:{user_personal_info[0]}\nکدملی:{user_personal_info[1]}')

            data_to_save = [user_personal_info[1],user_personal_info[0],context.user_data['user_choice_level_3'][0],context.user_data['user_choice_level_3'][1]]
            context.user_data['empty_cell'] = helper_funcs.empty_cell(sheet_visits)
            for j, value in enumerate(data_to_save, start=1):
                sheet_visits.cell(row=context.user_data['empty_cell'], column=j).value = value
            wb_visits.save('Userside bot\\visits.xlsx')

            # Resetting variables
            context.user_data['level'] = 0
            context.user_data['user_choice_level_1'] = None
            context.user_data['user_choice_level_2'] = None
            context.user_data['user_choice_level_3'] = None
            context.user_data['user_doctors'] = None
            context.user_data['empty_cell'] = None
            context.user_data['times'] = None
            
        else:
            await update.message.reply_text('❌ .پیام اشتباه! شماره تلفن باید با 09 شروع شود و کدملی هم می بایست 10 رقم باشد. همچنین اعداد باید انگلیسی وارد شده باشند.')

    else:
        await update.message.reply_text('❌ پیام اشتباه! ابتدا برای تهیه نوبت از دستور /visit  استفاده کنید.')


def main():
    application = Application.builder().token("7047332494:AAEsLSu5OJqCYQ1VBleQevBqEbOxQ_Sx_B0").build()

    # Command Handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("visit", visit_command))

    # Message Handlers
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, visit_process))
    
    
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()